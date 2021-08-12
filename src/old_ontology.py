import openpyxl
import json

# added ontology
Create_EV = {'template':'<arg1> created <arg2> using <arg3> at <arg4> place on <arg5> time', 'args':{'Creator':'arg1','Created_entity':'arg2','Instrument':'arg3','Place':'arg4','Time':'arg5'}}

class Ontology:
    def __init__(self, xlsx_path, case): # case : ['aida','kairos']
        self.data = self.load_xlsx(xlsx_path, case)
        self.events = self.data['events']
        self.relations = self.data['relations']
        self.entities = self.data['entities']
    
        # add custom events/relations
        self.events['Create.Create.Unspecified'] = Create_EV

    def load_xlsx(self, xlsx_path, case):
        wb = openpyxl.load_workbook(xlsx_path)
        ws_events = wb['events']
        ws_entities = wb['entities']
        ws_relations = wb['relations']
        data = {'events':{},'entities':{},'relations':{}}
        # load events:
        for row in ws_events.iter_rows(values_only = True):
            # get rid of None cells
            row = list(row)
            for i in range(len(row)):
                if row[i] is None:
                    row[i] = ''

            if row[0].strip() != 'AnnotIndexID':
                ev_type = row[1].strip()
                ev_subtype = row[3].strip()
                ev_subsubtype = row[5].strip()
                if case == 'aida':
                    if ev_subsubtype == 'n/a':
                        ev_type_str = '.'.join([ev_type,ev_subtype,'Unspecified'])
                    else:
                        ev_type_str = '.'.join([ev_type,ev_subtype,ev_subsubtype])
                elif case == 'kairos':
                    ev_type_str = '.'.join([ev_type,ev_subtype,ev_subsubtype])

                assert ev_type_str not in data['events']
                # template cell
                ev_template = row[8].strip()
                # arg cells
                if case == 'aida':
                    arg_roles = {row[9].strip():'arg1',row[12].strip():'arg2',row[15].strip():'arg3',row[18].strip():'arg4',row[21].strip():'arg5'}
                elif case == 'kairos':
                    arg_roles = {row[9].strip():'arg1',row[12].strip():'arg2',row[15].strip():'arg3',row[18].strip():'arg4',row[21].strip():'arg5',row[24].strip():'arg6'}

                data['events'][ev_type_str] = {'template':ev_template, 'args':arg_roles}
        # load relations:
        for row in ws_relations.iter_rows(values_only = True):
            
            row = list(row)
            for i in range(len(row)):
                if row[i] is None:
                    row[i] = ''
            
            if row[0].strip() != 'AnnotIndexID':
                re_type = row[1].strip()
                re_subtype = row[3].strip()
                re_subsubtype = row[5].strip()
                if case == 'aida':
                    if re_subsubtype == 'n/a':
                        re_type_str = '.'.join([re_type,re_subtype,'Unspecified'])
                    else:
                        re_type_str = '.'.join([re_type,re_subtype,re_subsubtype])
                elif case == 'kairos':
                    re_type_str = '.'.join([re_type,re_subtype,re_subsubtype])

                assert re_type_str not in data['relations']
                re_template = row[8].strip()
                re_args = {'Arg1':row[9].strip(), 'Arg2':row[12].strip()}
                data['relations'][re_type_str] = {'template':re_template, 'args':re_args}
                
        
        # TODO: entities
        return data
        
               